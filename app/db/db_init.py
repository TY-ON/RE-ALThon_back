import sqlite3
import time

conn = sqlite3.connect("database.db")
conn.execute(f"DROP TABLE user_data")
conn.execute("""CREATE TABLE user_data (
    userNum INTEGER PRIMARY KEY autoincrement,
    id text, 
    password text, 
    username text);""")

conn.execute(f"DROP TABLE product")
conn.execute("""CREATE TABLE product (
    prodNum INTEGER PRIMARY KEY autoincrement,
    name text, 
    company text,
    kind text, 
    
    price text,
    score text,
    
    content text,
    image text,
    useyn DEFAULT "y",
    regdate DATE, 
    prodSeq);""")

from openpyxl import load_workbook

load_wb = load_workbook("./db/input/input.xlsx", data_only=True)
sheet = load_wb.sheetnames
load_ws = load_wb[sheet[0]]

all_rows = load_ws.rows
is_first = True
for row in all_rows:
    if is_first:
        is_first = False
        continue
    conn.execute("INSERT INTO product \
                (name, company, kind, price, content, image, regdate) \
                     VALUES(?, ?, ?, ?, ?, ?, ?)", (row[0].value, row[3].value, row[2].value, f"{row[4].value//1000},{row[4].value%1000:03d}원", row[1].value, "row[5].value", time.time()))
conn.commit()
conn.close()